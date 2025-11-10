#!/usr/bin/env python3
"""
Export/Import SQLite database to/from Excel
Reads the processed_materials.db database and exports it to an Excel file,
or imports an Excel file back into the database
"""

import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook

# Configuration Constants
DB_FILE_NAME = "materials/processed_materials.db"
EXCEL_FILE_NAME = "materials/processed_materials.xlsx"
MODE = "import"



def export_database_to_excel(db_path, output_excel_path):
    """
    Export SQLite database to Excel file
    
    Args:
        db_path: Path to the SQLite database file
        output_excel_path: Path where the Excel file will be saved
    """
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # Get column names
        cursor.execute("SELECT * FROM materials")
        columns = [description[0] for description in cursor.description]
        
        # Get all data
        rows = cursor.fetchall()
        
        # Find and remove status column
        status_index = None
        if 'status' in columns:
            status_index = columns.index('status')
            columns.remove('status')
        
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Materials"
        
        # Write column headers
        ws.append(columns)
        
        # Write data rows (excluding status column)
        for row in rows:
            if status_index is not None:
                row_list = list(row)
                row_list.pop(status_index)
                ws.append(row_list)
            else:
                ws.append(row)
        
        # Auto-adjust column widths (no cap)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_excel_path)
        
        print(f"✓ Successfully exported {len(rows)} rows to {output_excel_path}")
        print(f"\nColumns exported:")
        for col in columns:
            print(f"  - {col}")
        
        print(f"\nPreview of first 5 rows:")
        for i, row in enumerate(rows[:5], 1):
            print(f"  Row {i}: {row}")
        
    except Exception as e:
        print(f"Error: {e}")
        raise
    
    finally:
        # Close the database connection
        conn.close()


def import_excel_to_database(excel_path, db_path):
    """
    Import Excel file to SQLite database
    
    Args:
        excel_path: Path to the Excel file
        db_path: Path to the SQLite database file
    """
    # Load the workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Get all rows from the worksheet
    rows = list(ws.values)
    
    if not rows:
        print("Error: Excel file is empty")
        return
    
    # First row contains column names
    excel_columns = list(rows[0])
    data_rows = rows[1:]
    
    # Connect to the SQLite database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    try:
        # Get the full column list from the database to include status
        cursor.execute("SELECT * FROM materials LIMIT 0")
        db_columns = [description[0] for description in cursor.description]
        
        # Find indices of key fields in Excel columns
        description_idx = excel_columns.index('description') if 'description' in excel_columns else None
        needed_idx = excel_columns.index('needed') if 'needed' in excel_columns else None
        key_concept_idx = excel_columns.index('key_concept') if 'key_concept' in excel_columns else None
        
        # Clear existing data from the table
        cursor.execute("DELETE FROM materials")
        
        # Add status column if not in Excel columns
        insert_columns = excel_columns.copy()
        if 'status' not in insert_columns:
            insert_columns.append('status')
        
        # Prepare the INSERT statement
        placeholders = ','.join(['?' for _ in insert_columns])
        column_names = ','.join(insert_columns)
        insert_query = f"INSERT INTO materials ({column_names}) VALUES ({placeholders})"
        
        # Process and insert rows with status
        rows_to_insert = []
        for row in data_rows:
            row_list = list(row)
            
            # Determine status based on three key fields
            description_filled = (description_idx is not None and 
                                row_list[description_idx] not in (None, '', ' '))
            needed_filled = (needed_idx is not None and 
                           row_list[needed_idx] is not None)
            key_concept_filled = (key_concept_idx is not None and 
                                row_list[key_concept_idx] not in (None, '', ' '))
            
            # Assign status "processed" if all three key fields are filled
            if description_filled and needed_filled and key_concept_filled:
                status = "processed"
            else:
                status = ""
            
            # Add status to row if not already present
            if 'status' not in excel_columns:
                row_list.append(status)
            
            rows_to_insert.append(tuple(row_list))
        
        # Insert all rows
        cursor.executemany(insert_query, rows_to_insert)
        
        # Commit the changes
        conn.commit()
        
        print(f"✓ Successfully imported {len(rows_to_insert)} rows from {excel_path}")
        print(f"\nColumns imported:")
        for col in insert_columns:
            print(f"  - {col}")
        
        print(f"\nPreview of first 5 rows:")
        for i, row in enumerate(rows_to_insert[:5], 1):
            print(f"  Row {i}: {row}")
        
    except Exception as e:
        conn.rollback()
        print(f"Error during import: {e}")
        raise
    
    finally:
        # Close the database connection
        conn.close()


if __name__ == "__main__":
    # Define paths using constants
    db_path = Path(DB_FILE_NAME)
    excel_path = Path(EXCEL_FILE_NAME)
    
    if MODE.lower() == "export":
        # Check if database exists
        if not db_path.exists():
            print(f"Error: Database file not found at {db_path}")
            exit(1)
        
        # Export to Excel
        export_database_to_excel(db_path, excel_path)
        print(f"\n✓ Export complete! File saved to: {excel_path}")
    
    elif MODE.lower() == "import":
        # Check if Excel file exists
        if not excel_path.exists():
            print(f"Error: Excel file not found at {excel_path}")
            exit(1)
        
        # Check if database exists
        if not db_path.exists():
            print(f"Error: Database file not found at {db_path}")
            exit(1)
        
        # Import from Excel
        import_excel_to_database(excel_path, db_path)
        print(f"\n✓ Import complete! Data loaded into: {db_path}")
    
    else:
        print(f"Error: Invalid MODE '{MODE}'. Must be 'export' or 'import'")
        exit(1)

